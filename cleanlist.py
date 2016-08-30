from openpyxl import load_workbook
from translate import trans


def xlfriendly(schedule, orig_xl):
    # Replaces days of the week with their coordinates
    for name in schedule:
        for shift in schedule[name]:
            if shift[0] in trans.keys():
                shift[0] = trans[shift[0]]

    # Load in Our Excel document
    wb = load_workbook(orig_xl)
    sh = wb.get_sheet_by_name('Sheet1')
    # Find the name in the Excel Document, replace it with Row value
    for i in range(1, 42):
        if sh.cell(row=i, column=1).value in schedule.keys():
            schedule[i] = schedule[sh.cell(row=i, column=1).value]
            del schedule[sh.cell(row=i, column=1).value]
    tbwschedule = schedule
    return tbwschedule
