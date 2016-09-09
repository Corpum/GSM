from openpyxl import load_workbook


class Spreadsheet():
    '''Creates an Excel spreadsheet containing information built in algo3'''
    def __init__(self, filename, startdate, tbwschedule, folder):
        self.folder = folder
        self.filename = filename
        self.createdoc()
        self.writeshifts(tbwschedule)
        self.writeoffdays()
        self.writedate(startdate)

    def createdoc(self):
        self.wb = load_workbook(self.filename)
        self.sh = self.wb.get_sheet_by_name('Sheet1')

    def writeshifts(self, tbwschedule):
        '''Writes information from Week() to Excel sheet'''
        for r in tbwschedule:
            for values in tbwschedule[r]:
                # Writes Shift start time
                self.sh.cell(row=r, column=values[0]).value = str(values[1][0])
                # Writes shift end time
                self.sh.cell(row=r, column=values[0]+1).value = str(
                        values[1][1])
                # Write in hours worked
                self.sh.cell(row=r, column=values[0]+2).value = values[1][2]

    def writeoffdays(self):
        ''' Adds in OFF days'''
        for j in range(8, 50):
            for i in range(4, 29, 3):
                if self.sh.cell(row=j, column=i).value in ('X', '#'):
                    self.sh.cell(row=j, column=i).value = 'OFF'
                    self.sh.cell(row=j, column=i + 1).value = ''
                    self.sh.cell(row=j, column=i + 2).value = 0

    def writedate(self, date):
        ''' Writes the beginning of week date to Excel'''
        self.sh.cell(row=4, column=4).value = date
        self.wb.save(self.folder + '/updated_sched3.xlsx')

if __name__ == '__main__':

    s = Spreadsheet()
