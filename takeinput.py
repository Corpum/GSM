from openpyxl import load_workbook
from translate import trans
import sqlite3 as lite
from tkinter.filedialog import askopenfilename


class Input():

    def __init__(self):
        self.trans2 = {v: k for k, v in trans.items()}
        self.info = {}
        self.id = 1
        self.position = 1
        self.load_doc()
        self.build_dict()
        self.init_db()
        self.build_db()
        self.insert()

    def load_doc(self):
        filename = askopenfilename()
        self.wb = load_workbook(filename)
        self.sh = self.wb.get_sheet_by_name('Sheet1')

    def build_dict(self):
        '''Search the lefthand column for names'''
        # Search through all names in column, for each name add in info.
        for r in range(1, 50):
            entry = str(self.sh.cell(row=r, column=1).value)
            if entry.isupper():
                # Find values for ID, position
                self.info[entry] = [self.id, entry, self.position]
                self.id += 1
                _type = self.sh.cell(row=r, column=6).value
                # Enter in position of each employee based on passed in
                # variables or lack thereof.
                if self.position == 1:
                    self.info[entry].append(0)
                elif _type == 'F':
                    self.info[entry].append(1)
                elif _type == 'B':
                    self.info[entry].append(2)
                else:
                    self.info[entry].append(3)
                for c in [4, 7, 10, 13, 16, 19, 22]:
                    availability = str(self.sh.cell(row=r, column=c).value)
                    if availability == 'X':
                        self.info[entry].append(self.trans2[c])
            elif entry == 'Management':
                self.position = 1

            elif entry == 'Full Time Store Associates':
                self.position = 2

            elif entry == 'Part Time Associates':
                self.position = 3

            else:
                continue

    def init_db(self):
        ''' Start a database'''
        self.db = lite.connect('stafflist.db')
        self.cur = self.db.cursor()

    def build_db(self):
        ''' Create the staff table in database'''
        self.cur.execute('CREATE TABLE staff ('
                         + 'id INTEGER NOT NULL PRIMARY KEY,'
                         + 'name VARCHAR(25) UNIQUE,'
                         + 'position INTEGER NOT NULL,'
                         + 'type INTEGER NOT NULL,'
                         + 'hours FLOAT DEFAULT 0,'
                         + 'maxhours FLOAT DEFAULT 0,'
                         + 'sunday INTEGER DEFAULT 1,'
                         + 'monday INTEGER DEFAULT 1,'
                         + 'tuesday INTEGER DEFAULT 1,'
                         + 'wednesday INTEGER DEFAULT 1,'
                         + 'thursday INTEGER DEFAULT 1,'
                         + 'friday INTEGER DEFAULT 1,'
                         + 'saturday INTEGER DEFAULT 1)')

    def insert(self):
        '''Insert the basic values and availability of each employee'''
        for emp in self.info:
            self.cur.execute('INSERT INTO staff(id, name, position, type)'
                             + 'VALUES(?, ?, ?, ?)', tuple(self.info[emp][:4]))
        for day in ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday',
                    'friday', 'saturday']:
            updatelist = [i for i in self.info if day in self.info[i]]
            for workers in updatelist:
                    self.cur.execute(('UPDATE staff SET {}=0 WHERE name=?')
                                     .format(day), (workers,))
        self.db.commit()
if __name__ == '__main__':

    i = Input()
